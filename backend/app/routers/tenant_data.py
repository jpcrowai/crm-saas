from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from typing import List, Optional, Dict, Any
import uuid
import json
from datetime import datetime, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from app.database import get_db
from app.deps import get_current_tenant_user, get_current_master
from app.models.schemas import Lead, LeadCreate, DashboardStats, TokenData, LeadHistory, LeadTask, Customer, CustomerCreate, TenantAdminStats
from app.models.sql_models import (
    Lead as SQLLead, 
    LeadHistory as SQLLeadHistory, 
    LeadTask as SQLLeadTask, 
    Customer as SQLCustomer, 
    Tenant as SQLTenant,
    Niche as SQLNiche,
    PipelineStage as SQLPipelineStage,
    User as SQLUser,
    FinanceEntry as SQLFinanceEntry
)

router = APIRouter(prefix="/tenant", tags=["tenant_data"])

# --- HELPERS ---

def sync_customer_from_lead(db: Session, tenant_id: uuid.UUID, lead: SQLLead):
    """Ensure a lead is also represented in the customers table."""
    # Try to match by lead_id first
    customer = db.query(SQLCustomer).filter(
        SQLCustomer.tenant_id == tenant_id,
        SQLCustomer.lead_id == lead.id
    ).first()
    
    if not customer:
        # Match by email or phone
        if lead.email:
            customer = db.query(SQLCustomer).filter(
                SQLCustomer.tenant_id == tenant_id,
                SQLCustomer.email == lead.email
            ).first()
        
        if not customer and lead.phone:
            # Simple phone matching (could be improved)
            customer = db.query(SQLCustomer).filter(
                SQLCustomer.tenant_id == tenant_id,
                SQLCustomer.phone == lead.phone
            ).first()

    cust_data = {
        "tenant_id": tenant_id,
        "name": lead.name,
        "email": lead.email,
        "phone": lead.phone,
        "lead_id": lead.id
    }

    if customer:
        for key, value in cust_data.items():
            setattr(customer, key, value)
    else:
        customer = SQLCustomer(**cust_data)
        db.add(customer)
    
    db.commit()

# --- ENDPOINTS ---

@router.get("/niche-config")
async def get_niche_config(
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    """Get niche configuration including pipeline stages for tenant's environment."""
    tenant = db.query(SQLTenant).filter(SQLTenant.id == current_user.tenant_id).first()
    if not tenant:
        return {"pipeline_stages": ["Novo", "Em Contato", "Agendado"]}

    # 1. Custom Stages
    stages = db.query(SQLPipelineStage).filter(
        SQLPipelineStage.tenant_id == tenant.id,
        SQLPipelineStage.active == True
    ).order_by(SQLPipelineStage.order_index).all()
    
    if stages:
        return {
            "niche_name": tenant.niche.name if tenant.niche else "Custom",
            "pipeline_stages": [s.name for s in stages]
        }

    # 2. Fallback to Niche Default if exists
    if tenant.niche:
        # Default stages for niche could be another table, 
        # but for now we'll return a static default or common ones
        return {
            "niche_name": tenant.niche.name,
            "pipeline_stages": ["Novo", "Em Contato", "Agendado"]
        }
    
    return {"pipeline_stages": ["Novo", "Em Contato", "Agendado"]}

@router.post("/pipeline-stages")
async def save_pipeline_stages(
    data: Dict[str, Any], 
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    stages_names = data.get("stages")
    if not isinstance(stages_names, list):
        raise HTTPException(status_code=400, detail="Stages must be a list")
    
    # Soft delete/Deactivate old stages
    db.query(SQLPipelineStage).filter(SQLPipelineStage.tenant_id == current_user.tenant_id).update({"active": False})
    
    # Add new ones
    for idx, name in enumerate(stages_names):
        new_stage = SQLPipelineStage(
            tenant_id=current_user.tenant_id,
            name=name,
            order_index=idx,
            active=True
        )
        db.add(new_stage)
    
    db.commit()
    return {"status": "success", "stages": stages_names}

@router.get("/leads", response_model=List[Lead])
async def get_leads(
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    leads = db.query(SQLLead).filter(SQLLead.tenant_id == current_user.tenant_id).order_by(desc(SQLLead.created_at)).all()
    return leads

@router.post("/leads", response_model=Lead)
async def create_lead(
    lead_in: Dict[str, Any], 
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    # Mapping fields from frontend (nome -> name, valor -> value, etc)
    new_lead = SQLLead(
        id=uuid.uuid4(),
        tenant_id=current_user.tenant_id,
        name=lead_in.get("nome") or lead_in.get("name") or "Lead sem nome",
        email=lead_in.get("email"),
        phone=lead_in.get("telefone") or lead_in.get("phone"),
        funil_stage=lead_in.get("funil_stage") or lead_in.get("status") or "new",
        value=lead_in.get("valor") or lead_in.get("value") or 0.0,
        origin=lead_in.get("origem") or lead_in.get("origin"),
        observations=lead_in.get("observacoes") or lead_in.get("observations"),
        responsible_user=lead_in.get("responsavel") or lead_in.get("responsible_user")
    )
    
    db.add(new_lead)
    db.flush()
    
    # Log initial history
    history = SQLLeadHistory(
        lead_id=new_lead.id,
        type="note",
        description="Lead criado no sistema",
        user_name=current_user.email
    )
    db.add(history)
    db.commit()
    db.refresh(new_lead)
    
    # Sync with customers
    sync_customer_from_lead(db, current_user.tenant_id, new_lead)
    
    return new_lead

@router.put("/leads/{lead_id}", response_model=Lead)
async def update_lead(
    lead_id: str, 
    lead_update: Dict[str, Any], 
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    lead = db.query(SQLLead).filter(SQLLead.id == lead_id, SQLLead.tenant_id == current_user.tenant_id).first()
    if not lead:
         raise HTTPException(status_code=404, detail="Lead not found")
         
    old_status = lead.funil_stage
    
    # Update fields
    if "nome" in lead_update: lead.name = lead_update["nome"]
    if "name" in lead_update: lead.name = lead_update["name"]
    if "email" in lead_update: lead.email = lead_update["email"]
    if "telefone" in lead_update: lead.phone = lead_update["telefone"]
    if "phone" in lead_update: lead.phone = lead_update["phone"]
    if "funil_stage" in lead_update: lead.funil_stage = lead_update["funil_stage"]
    if "status" in lead_update: lead.funil_stage = lead_update["status"]
    if "valor" in lead_update: lead.value = lead_update["valor"]
    if "value" in lead_update: lead.value = lead_update["value"]
    if "origem" in lead_update: lead.origin = lead_update["origem"]
    if "origin" in lead_update: lead.origin = lead_update["origin"]
    if "observacoes" in lead_update: lead.observations = lead_update["observacoes"]
    if "observations" in lead_update: lead.observations = lead_update["observations"]
    if "responsavel" in lead_update: lead.responsible_user = lead_update["responsavel"]
    
    # Log stage change
    if lead.funil_stage != old_status:
        history = SQLLeadHistory(
            lead_id=lead.id,
            type="stage_change",
            description=f"Mudança de etapa: {old_status} -> {lead.funil_stage}",
            user_name=current_user.email
        )
        db.add(history)
    
    db.commit()
    db.refresh(lead)
    
    # Sync with customers
    sync_customer_from_lead(db, current_user.tenant_id, lead)
    
    return lead

@router.get("/stats", response_model=DashboardStats)
async def get_stats(
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    tenant_id = current_user.tenant_id
    
    total_leads = db.query(SQLLead).filter(SQLLead.tenant_id == tenant_id).count()
    active_leads = db.query(SQLLead).filter(
        SQLLead.tenant_id == tenant_id, 
        SQLLead.funil_stage.in_(["new", "contacted", "Novo", "Em Contato"])
    ).count()
    converted_leads = db.query(SQLLead).filter(
        SQLLead.tenant_id == tenant_id, 
        SQLLead.funil_stage.in_(["converted", "Convertido"])
    ).count()
    
    revenue = db.query(func.sum(SQLLead.value)).filter(
        SQLLead.tenant_id == tenant_id, 
        SQLLead.funil_stage.in_(["converted", "Convertido"])
    ).scalar() or 0.0
    
    conversion_rate = (converted_leads / total_leads * 100) if total_leads > 0 else 0.0
    
    recent_leads = db.query(SQLLead).filter(SQLLead.tenant_id == tenant_id).order_by(desc(SQLLead.created_at)).limit(5).all()
    
    return DashboardStats(
        total_leads=total_leads,
        active_leads=active_leads,
        converted_leads=converted_leads,
        conversion_rate=round(conversion_rate, 2),
        total_revenue=float(revenue),
        recent_leads=recent_leads
    )

@router.get("/admin-stats", response_model=TenantAdminStats)
async def get_admin_stats(
    current_user: TokenData = Depends(get_current_master),
    db: Session = Depends(get_db)
):
    if not current_user.tenant_slug:
         raise HTTPException(status_code=400, detail="Not in tenant context")

    tenant = db.query(SQLTenant).filter(SQLTenant.slug == current_user.tenant_slug).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Environment not found")
        
    user_count = db.query(SQLUser).filter(SQLUser.tenant_id == tenant.id).count()
    admin_user = db.query(SQLUser).filter(SQLUser.tenant_id == tenant.id, SQLUser.role == "admin").first()

    return TenantAdminStats(
        nome=tenant.name,
        slug=tenant.slug,
        plan=tenant.plan_tier or "basic",
        payment_status=tenant.payment_status or "trial",
        ativo=tenant.active,
        user_count=user_count,
        admin_email=admin_user.email if admin_user else "N/A"
    )

@router.get("/revenue-chart")
async def get_revenue_chart(
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    # Aggregate converted leads by month for the last 6 months
    today = datetime.now()
    six_months_ago = today - timedelta(days=180)
    
    results = db.query(
        func.date_trunc('month', SQLLead.created_at).label('month'),
        func.sum(SQLLead.value).label('total')
    ).filter(
        SQLLead.tenant_id == current_user.tenant_id,
        SQLLead.status.in_(["converted", "Convertido"]),
        SQLLead.created_at >= six_months_ago
    ).group_by('month').order_by('month').all()
    
    return [
        {"name": r.month.strftime("%b %y"), "value": float(r.total)}
        for r in results
    ]

@router.get("/leads/{lead_id}/history", response_model=List[LeadHistory])
async def get_lead_history(
    lead_id: str, 
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    return db.query(SQLLeadHistory).filter(SQLLeadHistory.lead_id == lead_id).order_by(desc(SQLLeadHistory.created_at)).all()

@router.post("/leads/{lead_id}/history", response_model=LeadHistory)
async def add_lead_history(
    lead_id: str, 
    item_in: Dict[str, Any], 
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    new_item = SQLLeadHistory(
        id=uuid.uuid4(),
        lead_id=lead_id,
        type=item_in.get("type", "note"),
        description=item_in.get("description", ""),
        user_name=current_user.email
    )
    db.add(new_item)
    db.commit()
    db.refresh(new_item)
    return new_item

@router.get("/leads/{lead_id}/tasks", response_model=List[LeadTask])
async def get_lead_tasks(
    lead_id: str, 
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    return db.query(SQLLeadTask).filter(SQLLeadTask.lead_id == lead_id).order_by(SQLLeadTask.due_date).all()

@router.post("/leads/{lead_id}/tasks", response_model=LeadTask)
async def create_lead_task(
    lead_id: str, 
    task_in: Dict[str, Any], 
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    new_task = SQLLeadTask(
        id=uuid.uuid4(),
        lead_id=lead_id,
        title=task_in.get("title", ""),
        due_date=datetime.fromisoformat(task_in["due_date"]) if task_in.get("due_date") else None,
        responsible_user=task_in.get("responsible_user", current_user.email),
        status="pending"
    )
    db.add(new_task)
    db.commit()
    db.refresh(new_task)
    return new_task

@router.put("/tasks/{task_id}", response_model=LeadTask)
async def update_task(
    task_id: str, 
    task_update: Dict[str, Any], 
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    task = db.query(SQLLeadTask).filter(SQLLeadTask.id == task_id).first()
    if not task: raise HTTPException(status_code=404)
    
    if "title" in task_update: task.title = task_update["title"]
    if "status" in task_update: task.status = task_update["status"]
    if "due_date" in task_update: task.due_date = datetime.fromisoformat(task_update["due_date"]) if task_update["due_date"] else None
    if "responsible_user" in task_update: task.responsible_user = task_update["responsible_user"]
    
    db.commit()
    db.refresh(task)
    return task

@router.post("/leads/import-excel")
async def import_leads_excel(
    file: UploadFile = File(...), 
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    from openpyxl import load_workbook
    from io import BytesIO
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser Excel")
        
    try:
        contents = await file.read()
        wb = load_workbook(BytesIO(contents), data_only=True)
        ws = wb.active
        
        # Get headers
        headers = [str(cell.value).lower().strip() for cell in ws[1] if cell.value is not None]
        
        mapping = {
            "nome": ["nome", "lead", "cliente", "name"],
            "telefone": ["telefone", "celular", "whatsapp", "phone", "tel"],
            "email": ["email", "e-mail", "mail"],
            "origem": ["origem", "source", "canal"],
            "responsavel": ["responsavel", "dono", "owner", "responsible"],
            "status": ["status", "etapa", "stage"],
            "valor": ["valor", "oportunidade", "value", "price"],
            "observacoes": ["observacoes", "notas", "obs", "notes"]
        }
        
        def find_col_idx(possible_names):
            for idx, h in enumerate(headers):
                if h in possible_names: return idx
            return None

        col_mapping = {custom: find_col_idx(standards) for custom, standards in mapping.items()}
        stats = {"created": 0, "updated": 0, "errors": 0}
        
        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                row_data = {}
                for key, idx in col_mapping.items():
                    if idx is not None:
                        val = row[idx]
                        row_data[key] = str(val) if val is not None else ""
                    else:
                        row_data[key] = ""
                
                if not row_data["nome"]: continue

                # Try match existing
                lead = db.query(SQLLead).filter(
                    SQLLead.tenant_id == current_user.tenant_id,
                    SQLLead.name == row_data["nome"]
                ).first()
                
                if lead:
                    lead.phone = row_data["telefone"]
                    lead.email = row_data["email"]
                    lead.origin = row_data["origem"]
                    lead.status = row_data["status"] or lead.status
                    stats["updated"] += 1
                else:
                    lead = SQLLead(
                        tenant_id=current_user.tenant_id,
                        name=row_data["nome"],
                        phone=row_data["telefone"],
                        email=row_data["email"],
                        origin=row_data["origem"],
                        status=row_data["status"] or "new",
                        value=float(row_data["valor"]) if row_data["valor"] else 0.0,
                        observations=row_data["observacoes"],
                        responsible_user=row_data["responsavel"]
                    )
                    db.add(lead)
                    stats["created"] += 1
                
                db.flush()
                # history
                db.add(SQLLeadHistory(lead_id=lead.id, type="note", description="Importado via Excel", user_name=current_user.email))
                sync_customer_from_lead(db, current_user.tenant_id, lead)
            except Exception as e:
                print(f"Error importing row: {e}")
                stats["errors"] += 1
        
        db.commit()
        return stats
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/customers", response_model=List[Customer])
async def get_customers(
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    return db.query(SQLCustomer).filter(SQLCustomer.tenant_id == current_user.tenant_id).all()

@router.post("/customers", response_model=Customer)
async def create_customer(
    customer_in: Dict[str, Any], 
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    new_cust = SQLCustomer(
        tenant_id=current_user.tenant_id,
        name=customer_in.get("name") or customer_in.get("nome"),
        email=customer_in.get("email"),
        phone=customer_in.get("phone") or customer_in.get("telefone"),
        document=customer_in.get("document"),
        address=customer_in.get("address"),
        lead_id=customer_in.get("lead_id")
    )
    db.add(new_cust)
    db.commit()
    db.refresh(new_cust)
    return new_cust

@router.delete("/customers/{cust_id}")
async def delete_customer(
    cust_id: str, 
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    cust = db.query(SQLCustomer).filter(SQLCustomer.id == cust_id, SQLCustomer.tenant_id == current_user.tenant_id).first()
    if not cust: raise HTTPException(status_code=404)
    db.delete(cust)
    db.commit()
    return {"message": "Deletado"}

@router.get("/reports")
async def get_reports(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    tenant_id = current_user.tenant_id
    
    # 1. Total Leads (Active in period)
    leads_query = db.query(SQLLead).filter(SQLLead.tenant_id == tenant_id)
    if start_date:
        leads_query = leads_query.filter(SQLLead.created_at >= start_date)
    if end_date:
        leads_query = leads_query.filter(SQLLead.created_at <= end_date)
    total_leads = leads_query.count()
    
    # 2. Financial Aggregate
    # Base query for finances
    base_finances = db.query(SQLFinanceEntry.amount).filter(
        SQLFinanceEntry.tenant_id == tenant_id,
        SQLFinanceEntry.status == "pago"
    )
    
    if start_date:
        base_finances = base_finances.filter(SQLFinanceEntry.due_date >= start_date)
    if end_date:
        base_finances = base_finances.filter(SQLFinanceEntry.due_date <= end_date)
        
    revenue_paid = base_finances.filter(SQLFinanceEntry.type == "receita").with_entities(func.sum(SQLFinanceEntry.amount)).scalar() or 0.0
    expenses_paid = base_finances.filter(SQLFinanceEntry.type == "despesa").with_entities(func.sum(SQLFinanceEntry.amount)).scalar() or 0.0
    
    # 3. Customer Ranking
    ranking_query = db.query(
        SQLCustomer.name,
        func.sum(SQLFinanceEntry.amount).label('total')
    ).join(
        SQLFinanceEntry, SQLFinanceEntry.customer_id == SQLCustomer.id
    ).filter(
        SQLCustomer.tenant_id == tenant_id,
        SQLFinanceEntry.type == "receita",
        SQLFinanceEntry.status == "pago"
    )
    
    if start_date:
        ranking_query = ranking_query.filter(SQLFinanceEntry.due_date >= start_date)
    if end_date:
        ranking_query = ranking_query.filter(SQLFinanceEntry.due_date <= end_date)
        
    ranking_raw = ranking_query.group_by(SQLCustomer.name).order_by(desc('total')).limit(5).all()
    
    ranking = [{"customer_name": r.name, "total_revenue": float(r.total)} for r in ranking_raw]

    return {
        "total_leads": total_leads,
        "total_revenue": float(revenue_paid),
        "total_expenses": float(expenses_paid),
        "customer_ranking": ranking
    }
