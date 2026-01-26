import os
import openpyxl
import json
from typing import List, Dict, Any

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
EXCEL_DIR = os.path.join(BASE_DIR, "excel")

if not os.path.exists(EXCEL_DIR):
    os.makedirs(EXCEL_DIR)

def _get_file_path(file_name: str) -> str:
    if not file_name.endswith(".xlsx"):
        file_name += ".xlsx"
    return os.path.join(EXCEL_DIR, file_name)

def read_sheet(file_name: str, sheet_name: str) -> List[Dict[str, Any]]:
    path = _get_file_path(file_name)
    if not os.path.exists(path):
        return []
    
    try:
        workbook = openpyxl.load_workbook(path)
        if sheet_name not in workbook.sheetnames:
            return []
        
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows:
            return []
            
        headers = rows[0]
        data_rows = rows[1:]
        
        result = []
        for row in data_rows:
            row_data = {}
            for i, header in enumerate(headers):
                if header:
                    val = row[i] if i < len(row) else None
                    # Try to parse JSON if it looks like a list or dict
                    if isinstance(val, str):
                        stripped = val.strip()
                        if (stripped.startswith("[") and stripped.endswith("]")) or (stripped.startswith("{") and stripped.endswith("}")):
                            try:
                                val = json.loads(stripped)
                            except:
                                pass
                    row_data[str(header)] = val
            result.append(row_data)
            
        return result
    except Exception as e:
        print(f"Error reading excel {file_name}/{sheet_name}: {e}")
        return []

def write_sheet(file_name: str, sheet_name: str, rows: List[Dict[str, Any]]) -> None:
    path = _get_file_path(file_name)
    
    if os.path.exists(path):
        workbook = openpyxl.load_workbook(path)
    else:
        workbook = openpyxl.Workbook()
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]
            
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    
    sheet = workbook.create_sheet(sheet_name)
    
    if not rows:
        workbook.save(path)
        return

    all_keys = set()
    for r in rows:
        all_keys.update(r.keys())
    
    headers = list(all_keys)
    sheet.append(headers)
    
    for row_dict in rows:
        row_val = []
        for h in headers:
            val = row_dict.get(h)
            # Serialize complex types to JSON strings
            if isinstance(val, (list, dict)):
                val = json.dumps(val)
            row_val.append(val)
        sheet.append(row_val)
        
    workbook.save(path)

def init_excel_file(file_name: str, sheets_columns: Dict[str, List[str]]):
    path = _get_file_path(file_name)
    if os.path.exists(path):
        return

    workbook = openpyxl.Workbook()
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]
        
    for sheet_name, columns in sheets_columns.items():
        ws = workbook.create_sheet(sheet_name)
        ws.append(columns)
        
    workbook.save(path)
