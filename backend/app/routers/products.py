from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from typing import List, Optional, Dict, Any
from openpyxl import Workbook, load_workbook
from io import BytesIO
from datetime import datetime
from pydantic import BaseModel
from sqlalchemy.orm import Session
from app.database import get_db
from app.deps import get_current_tenant_user
from app.models.schemas import TokenData
from app.models.sql_models import Product as SQLProduct

router = APIRouter(prefix="/tenant", tags=["products"])

class Product(BaseModel):
    id: Optional[str] = None
    sku: str
    name: str
    description: Optional[str] = ""
    price: float
    category: str = ""
    duration: int = 30
    unit: str = "unit"
    status: str = "Active"
    created_at: Optional[str] = None
    updated_at: Optional[str] = None

    class Config:
        from_attributes = True

def clean_price(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    if isinstance(val, str):
        clean = val.replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
        try:
            return float(clean)
        except:
            return 0.0
    return 0.0

@router.get("/products", response_model=List[Product])
async def get_products(
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    products = db.query(SQLProduct).filter(
        SQLProduct.tenant_id == current_user.tenant_id
    ).all()
    
    result = []
    for p in products:
        result.append({
            "id": str(p.id),
            "sku": p.sku if p.sku else "",
            "name": p.name,
            "description": p.description or "",
            "price": float(p.price),
            "category": "Service" if p.type == "service" else "Product",
            "duration": p.duration_minutes or 30,
            "unit": "unit",
            "status": "Active" if p.active else "Inactive",
            "created_at": p.created_at.isoformat() if p.created_at else None
        })
    
    return result

@router.post("/products", response_model=Product)
async def create_product(
    product: Product,
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    if product.sku:
        existing = db.query(SQLProduct).filter(
            SQLProduct.tenant_id == current_user.tenant_id,
            SQLProduct.sku == product.sku
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail=f"Product with code {product.sku} already exists")
    
    p_type = "service" if product.category.lower() == "service" else "product"

    new_product = SQLProduct(
        tenant_id=current_user.tenant_id,
        sku=product.sku,
        name=product.name,
        description=product.description,
        price=product.price,
        type=p_type,
        duration_minutes=product.duration,
        active=(product.status == "Active")
    )
    
    db.add(new_product)
    db.commit()
    db.refresh(new_product)
    
    return Product(
        id=str(new_product.id),
        sku=new_product.sku or "",
        name=new_product.name,
        description=new_product.description or "",
        price=float(new_product.price),
        category=product.category,
        duration=new_product.duration_minutes or 30,
        unit=product.unit,
        status="Active" if new_product.active else "Inactive",
        created_at=new_product.created_at.isoformat()
    )

@router.put("/products/{product_id}", response_model=Product)
async def update_product(
    product_id: str,
    product_in: Dict[str, Any],
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    product = db.query(SQLProduct).filter(
        SQLProduct.id == product_id,
        SQLProduct.tenant_id == current_user.tenant_id
    ).first()
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    if "sku" in product_in:
        product.sku = product_in["sku"]
    if "name" in product_in:
        product.name = product_in["name"]
    if "description" in product_in:
        product.description = product_in["description"]
    if "price" in product_in:
        product.price = clean_price(product_in["price"])
    if "status" in product_in:
        product.active = (product_in["status"] == "Active")
    if "category" in product_in:
        product.type = "service" if product_in["category"].lower() == "service" else "product"
    if "duration" in product_in:
        try:
            product.duration_minutes = int(product_in["duration"])
        except: pass
    
    db.commit()
    db.refresh(product)
    
    return Product(
        id=str(product.id),
        sku=product.sku or "",
        name=product.name,
        description=product.description or "",
        price=float(product.price),
        category="Service" if product.type == "service" else "Product",
        duration=product.duration_minutes or 30,
        unit=product_in.get("unit", "unit"),
        status="Active" if product.active else "Inactive",
        created_at=product.created_at.isoformat()
    )

@router.delete("/products/{product_id}")
async def delete_product(
    product_id: str,
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    product = db.query(SQLProduct).filter(
        SQLProduct.id == product_id,
        SQLProduct.tenant_id == current_user.tenant_id
    ).first()
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    product.active = False
    db.commit()
    
    return {"message": "Product deactivated successfully"}

@router.get("/products/template")
async def get_product_template(
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    products = db.query(SQLProduct).filter(
        SQLProduct.tenant_id == current_user.tenant_id
    ).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"
    
    headers = ["Código", "Nome", "Descrição", "Preço Unitário", "Tipo", "Duração (min)", "Categoria", "Unidade", "Status"]
    ws.append(headers)
    
    for p in products:
        ws.append([
            p.sku or "",
            p.name,
            p.description or "",
            float(p.price),
            "Serviço" if p.type == "service" else "Produto",
            p.duration_minutes or 30,
            "",
            "unit",
            "Ativo" if p.active else "Inativo"
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"modelo_produtos_{current_user.tenant_slug}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/products/import")
async def import_products(
    file: UploadFile = File(...),
    current_user: TokenData = Depends(get_current_tenant_user),
    db: Session = Depends(get_db)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be Excel")
        
    try:
        contents = await file.read()
        wb = load_workbook(BytesIO(contents), data_only=True)
        ws = wb.active
        
        # Get headers
        headers = [str(cell.value).lower().strip() for cell in ws[1] if cell.value is not None]
        
        mapping = {
            "sku": ["código", "codigo", "code", "id", "sku"],
            "name": ["nome", "name", "produto", "serviço", "servico"],
            "description": ["descrição", "descricao", "description", "obs"],
            "price": ["preço unitário", "preco unitario", "preço", "valor", "price"],
            "status": ["status", "ativo", "situacao"],
            "type": ["tipo", "type", "categoria"],
            "duration": ["duração", "duracao", "duration", "duração (min)"]
        }
        
        def find_col_idx(possible_names):
            for idx, h in enumerate(headers):
                if h in possible_names:
                    return idx
            return None

        col_mapping = {key: find_col_idx(standards) for key, standards in mapping.items()}
        
        if col_mapping["name"] is None and col_mapping["sku"] is None:
            raise HTTPException(status_code=400, detail="Column 'Nome' or 'Código' not found in Excel")

        stats = {"created": 0, "updated": 0, "errors": 0}
        
        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                name = ""
                if col_mapping["name"] is not None:
                    val = row[col_mapping["name"]]
                    name = str(val).strip() if val is not None else ""
                
                sku = ""
                if col_mapping["sku"] is not None:
                    val_sku = row[col_mapping["sku"]]
                    sku = str(val_sku).strip() if val_sku is not None else ""
                    
                if not name and sku:
                    name = sku
                
                if not name: continue
                
                existing = None
                if sku:
                    existing = db.query(SQLProduct).filter(
                        SQLProduct.tenant_id == current_user.tenant_id,
                        SQLProduct.sku == sku
                    ).first()
                
                if not existing:
                    existing = db.query(SQLProduct).filter(
                        SQLProduct.tenant_id == current_user.tenant_id,
                        SQLProduct.name == name
                    ).first()
                
                price = 0.0
                if col_mapping["price"] is not None:
                    val_price = row[col_mapping["price"]]
                    if val_price is not None:
                        price = clean_price(val_price)
                
                description = ""
                if col_mapping["description"] is not None:
                    val_desc = row[col_mapping["description"]]
                    description = str(val_desc) if val_desc is not None else ""
                
                active = True
                if col_mapping["status"] is not None:
                    val_stat = row[col_mapping["status"]]
                    status_val = str(val_stat).lower() if val_stat is not None else ""
                    active = status_val in ["active", "ativo", "ativa", "sim", "yes", "true", "1"]

                p_type = "product"
                if col_mapping["type"] is not None:
                    val_type = str(row[col_mapping["type"]]).lower() if row[col_mapping["type"]] is not None else ""
                    if "serv" in val_type:
                        p_type = "service"

                duration = 30
                if col_mapping["duration"] is not None:
                    try:
                        val_dur = row[col_mapping["duration"]]
                        if val_dur is not None:
                            duration = int(float(val_dur))
                    except: pass
                
                if existing:
                    if sku: existing.sku = sku
                    existing.name = name
                    existing.description = description
                    existing.price = price
                    existing.active = active
                    existing.type = p_type
                    existing.duration_minutes = duration
                    stats["updated"] += 1
                else:
                    new_product = SQLProduct(
                        tenant_id=current_user.tenant_id,
                        sku=sku,
                        name=name,
                        description=description,
                        price=price,
                        active=active,
                        type=p_type,
                        duration_minutes=duration
                    )
                    db.add(new_product)
                    stats["created"] += 1
                    
            except Exception as e:
                print(f"Error importing row: {e}")
                stats["errors"] += 1
                
        db.commit()
        
        return {
            "message": "Product import finished",
            "created": stats["created"],
            "updated": stats["updated"],
            "errors": stats["errors"]
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
